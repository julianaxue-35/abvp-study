#!/usr/bin/env python3
"""Extract new 2021-2024 folder articles, dedup vs catalog, emit worklist + report."""
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

TOOLS = Path(__file__).resolve().parent
sys.path.insert(0, str(TOOLS))
from lib_extract import (norm_title, find_cols, detect_year, author_lastname,
                         is_dup, fuzzy_matches)
from lib_catalog import load_catalog

SS = Path("/Users/jxue/Library/CloudStorage/Dropbox/work docs/protocols in interactive dashboards/ABVP exam hub/2026 ABVP shelter medicine study/JOURNAL REVIEW/SPREADSHEETS")
ACTIVE = ["2025 Diplomate publications spreadsheet.xlsx",
          "_Behaviour and Welfare Articles Board Review 2024.xlsx",
          "_Infectious Disease articles - 2024.xlsx",
          "_Misc Spreadsheet Study Guide - Boards 2024.xlsx"]

# sheet name (lower) -> (domain, subdomain_page). Stage-0 seed; the agent refines.
SHEET_MAP = {
    "cirdc": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "dermatophytosis": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "cpv": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "fpv": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "fiv.felv": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "feline uri": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "sars-cov-2": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "fip.fecov": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "other": ("Physical Health of Animal", "physical-health/infectious_disease_hub.html"),
    "heartworm": ("Physical Health of Animal", "physical-health/parasites_hub.html"),
    "parasitology": ("Physical Health of Animal", "physical-health/parasites_hub.html"),
    "anesthesiaanalgesia": ("Physical Health of Animal", "physical-health/surgery_anesthesia_hub.html"),
    "spayneuter surgery": ("Companion Animal Homelessness", "companion-animal-homelessness/spay_neuter_hub_2.html"),
    "agebenefits of spayneuter": ("Companion Animal Homelessness", "companion-animal-homelessness/spay_neuter_hub_2.html"),
    "snrtnr": ("Companion Animal Homelessness", "companion-animal-homelessness/nonsurgical_sterilization_hub.html"),
    "outreach for owned animals": ("Companion Animal Homelessness", "companion-animal-homelessness/access_vet_care_hub.html"),
    "transport": ("Companion Animal Homelessness", "companion-animal-homelessness/animal_transport_relocation_hub.html"),
    "disaster": ("Companion Animal Homelessness", "companion-animal-homelessness/disaster_emergency_hub.html"),
    "mgmt stats design sanitation": ("Shelter Management", "shelter-management/05_data_analysis.html"),
    "mental health": ("Shelter Management", "shelter-management/06_mental_health.html"),
    "forensicscrueltyhoarding": ("Community and Public Health", "community-public-health/01_animal_cruelty.html"),
    "public health  one health  zoon": ("Community and Public Health", "community-public-health/02_zoonotic_disease.html"),
    "k9 behaviour & training": ("Behavioral Health", "behavioral-health/14_training_bmod_playgroups.html"),
    "k9 behaviour assessments": ("Behavioral Health", "behavioral-health/03_assessment_decision_making.html"),
    "fel behaviour & training": ("Behavioral Health", "behavioral-health/14_training_bmod_playgroups.html"),
    "psychopharm & pheromones": ("Behavioral Health", "behavioral-health/09_behaviour_medications.html"),
    "welfarehousingc4c": ("Behavioral Health", "behavioral-health/07_facility_environment.html"),
    "intakeoutcomelos": ("Shelter Management", "shelter-management/05_data_analysis.html"),
    "pediatrics": ("Physical Health of Animal", "physical-health/medical_health_hub.html"),
    "small mammal, exotic, farm": ("Physical Health of Animal", "physical-health/other-animals_hub.html"),
    "general medicine": ("Physical Health of Animal", "physical-health/medical_health_hub.html"),
    "misc": ("Shelter Management", "shelter-management/03_management_leadership.html"),
    "diplomate publications": ("Companion Animal Homelessness", "companion-animal-homelessness/access_vet_care_hub.html"),
    "abvp journal club articles": ("Companion Animal Homelessness", "companion-animal-homelessness/access_vet_care_hub.html"),
}
DEFAULT_MAP = ("Shelter Management", "shelter-management/03_management_leadership.html")


def build_existing_index(catalog):
    idx = {"titles": set(), "author_year": set(), "by_ay": {}}
    for r in catalog:
        n = norm_title(r["title"])
        idx["titles"].add(n)
        al = author_lastname(r.get("authors", ""))
        yr = r.get("year")
        idx["author_year"].add((al, yr))
        idx["by_ay"].setdefault((al, yr), []).append(n)
    return idx


def main():
    catalog = load_catalog(str(TOOLS / "journal-catalog.json"))
    existing = build_existing_index(catalog)
    existing_norms = list(existing["titles"])
    staged, dropped, nearlist, seen = [], [], [], set()
    for fname in ACTIVE:
        wb = openpyxl.load_workbook(SS / fname, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            cols = find_cols(rows[0])
            if "title" not in cols or "abstract" not in cols:
                continue
            dom, page = SHEET_MAP.get(sn.strip().lower(), DEFAULT_MAP)
            for r in rows[1:]:
                tc = cols["title"]
                if tc >= len(r) or not r[tc] or not str(r[tc]).strip():
                    continue
                title = str(r[tc]).strip()
                year = detect_year(list(r), cols)
                if not year or not (2021 <= year <= 2024):
                    continue
                ab = r[cols["abstract"]] if cols["abstract"] < len(r) else None
                abstract = str(ab).strip() if ab else ""
                if len(abstract) < 40:
                    continue  # need a real abstract to write MCQs
                authors = (str(r[cols["author"]]).strip()
                           if cols.get("author") is not None and cols["author"] < len(r) and r[cols["author"]] else "")
                journal = (str(r[cols["journal"]]).strip()
                           if cols.get("journal") is not None and cols["journal"] < len(r) and r[cols["journal"]] else "")
                if cols.get("combined_author_journal"):
                    journal = ""  # author col holds both; agent refines
                kp = (str(r[cols["keypoints"]]).strip()
                      if cols.get("keypoints") is not None and cols["keypoints"] < len(r) and r[cols["keypoints"]] else "")
                n = norm_title(title)
                if n in seen:
                    continue
                al = author_lastname(authors)
                if is_dup(n, al, year, existing):
                    dropped.append({"title": title, "year": year, "sheet": sn})
                    continue
                fm = fuzzy_matches(n, existing_norms)
                if fm:
                    nearlist.append({"title": title, "year": year, "sheet": sn,
                                     "best": fm[0][0], "score": fm[0][1]})
                seen.add(n)
                staged.append({"title": title, "authors": authors, "journal": journal,
                               "year": year, "abstract": abstract, "key_points": kp,
                               "source_sheet": sn, "suggested_domain": dom,
                               "suggested_page": page, "source": "folder",
                               "abstract_origin": "spreadsheet", "mcqs": []})
        wb.close()
    (TOOLS / "folder_new_articles.json").write_text(
        json.dumps(staged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    by_year = Counter(s["year"] for s in staged)
    by_dom = Counter(s["suggested_domain"] for s in staged)
    lines = [
        "# Folder dedup report\n",
        f"- Staged NEW articles: **{len(staged)}**",
        f"- Exact duplicates dropped: {len(dropped)}",
        f"- Near-duplicates to eyeball: {len(nearlist)}\n",
        "## By year\n" + "\n".join(f"- {k}: {v}" for k, v in sorted(by_year.items())),
        "\n## By suggested domain\n" + "\n".join(f"- {k}: {v}" for k, v in sorted(by_dom.items())),
        "\n## Near-duplicates (review these)\n" + (
            "\n".join(f"- [{x['score']}] {x['year']} {x['title'][:80]}  ⟷  {x['best'][:50]}"
                      for x in nearlist) or "_none_"),
    ]
    (TOOLS / "folder_dedup_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Near-dups"
    ws.append(["Score", "Year", "Title", "Sheet", "Closest existing (normalized)"])
    for x in nearlist:
        ws.append([x["score"], x["year"], x["title"], x["sheet"], x["best"]])
    ws2 = out.create_sheet("Dropped exact dups")
    ws2.append(["Year", "Title", "Sheet"])
    for x in dropped:
        ws2.append([x["year"], x["title"], x["sheet"]])
    out.save(TOOLS / "folder_dedup_report.xlsx")
    print(f"staged={len(staged)} dropped={len(dropped)} near={len(nearlist)}")
    print("by_year:", dict(sorted(by_year.items())))


if __name__ == "__main__":
    main()
