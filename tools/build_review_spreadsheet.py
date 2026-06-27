#!/usr/bin/env python3
"""Build the 2025-2026 review spreadsheet (4 tabs, format per 2025 Diplomate sheet).

Reads tools/discovery_2025_2026/cand_deduped.json (deduped vs the hub) and writes
2025-2026-shelter-med-articles.xlsx at the repo root with four tabs:
Diplomate Publications | Misc | Behaviour | Infectious Disease.
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
TABS = ["Diplomate Publications", "Misc", "Behaviour", "Infectious Disease"]
HEADERS = ["Diplomate", "Journal", "Year", "Author", "Title", "Abstract", "Takeaways", "Source"]


def route(tab):
    t = (tab or "").strip().lower()
    if "diplomate" in t:
        return "Diplomate Publications"
    if "behav" in t:
        return "Behaviour"
    if "infect" in t:
        return "Infectious Disease"
    return "Misc"


def main():
    src = TOOLS / "discovery_2025_2026" / "cand_deduped.json"
    cands = json.loads(src.read_text(encoding="utf-8"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    counts = {}
    for tab in TABS:
        ws = wb.create_sheet(tab[:31])
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
        rows = [r for r in cands if route(r.get("topic_tab")) == tab]
        rows.sort(key=lambda r: (-int(r.get("year", 0) or 0), str(r.get("title", "")).lower()))
        for r in rows:
            ws.append([r.get("diplomate", ""), r.get("journal", ""), r.get("year", ""),
                       r.get("authors", ""), r.get("title", ""), r.get("abstract", ""),
                       r.get("takeaways", ""), r.get("source_url", "")])
        widths = [20, 18, 6, 26, 48, 80, 46, 34]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        counts[tab] = len(rows)
    out = ROOT / "2025-2026-shelter-med-articles.xlsx"
    wb.save(out)
    print(f"wrote {out.name}")
    for tab in TABS:
        print(f"  {tab}: {counts[tab]} articles")
    print(f"  TOTAL: {sum(counts.values())}")


if __name__ == "__main__":
    main()
